import uuid
import datetime
import re
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, distinct
from collections import Counter
from typing import Optional, List, Dict, Any
import asyncio
import os
import json
from openai import OpenAI
import base64
import fitz
import csv
import io
import openpyxl

import models
import schemas
from event_bus import global_event_bus, SystemEvent

# Import croniter for scheduled insights. This is a new dependency.
# Add 'croniter' to your requirements.txt
from croniter import croniter
from services.redis_middleware import redis_client, rate_limit_script
from services.maintenance_utils import log_maintenance_task
from services.asset_cache import AssetCache
from services.integration_dispatcher import IntegrationDispatcher

class AIService:
    """
    Service layer for AI/ML features, including data logging and model inference.
    """

    def _log_task(self, db: Session, task_name: str, status: str, triggered_by: str, summary: dict = None, details: str = None, duration_ms: int = None):
        """Helper method to log the execution of a maintenance task."""
        log_maintenance_task(db, task_name, status, triggered_by, summary, details, duration_ms)

    def log_user_interaction(self, db: Session, user_id: str, event_data: schemas.UserInteractionEventCreate) -> models.UserInteractionEvent:
        """
        Logs a user interaction event to the database for future AI/ML processing.
        """
        new_event = models.UserInteractionEvent(
            event_id=f"INTERACTION-{uuid.uuid4().hex}",
            user_id=user_id,
            timestamp=datetime.datetime.utcnow().isoformat(),
            **event_data.dict()
        )
        
        db.add(new_event)
        db.commit()
        db.refresh(new_event)
        
        return new_event

    def get_user_interaction_summary(self, db: Session, user_id: str, limit: int = 20) -> dict:
        """
        Retrieves a summary of a user's recent interactions.
        """
        interaction_query = db.query(models.UserInteractionEvent).filter(models.UserInteractionEvent.user_id == user_id)
        
        total_count = interaction_query.count()
        
        recent_interactions = interaction_query.order_by(desc(models.UserInteractionEvent.timestamp)).limit(limit).all()
        
        return {
            "user_id": user_id,
            "total_interactions": total_count,
            "recent_interactions": recent_interactions
        }

    def generate_predictive_insight(self, db: Session, user_id: str, current_event_type: str, current_target_id: Optional[str], history_limit: int = 1000) -> dict:
        """
        Analyzes a user's recent interaction history to predict the most likely next action.
        This is a simple implementation of Behavioural AI.
        """
        # Fetch the most recent interactions for the user, ordered reverse-chronologically
        recent_interactions = db.query(models.UserInteractionEvent).filter(
            models.UserInteractionEvent.user_id == user_id
        ).order_by(desc(models.UserInteractionEvent.timestamp)).limit(history_limit).all()

        # Reverse the list to get chronological order for pattern analysis
        all_interactions = list(reversed(recent_interactions))

        # --- LAYER 2: LLM Integration for Advanced Predictive Forecasting ---
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if openai_api_key and len(all_interactions) >= 3:
            try:
                client = OpenAI(api_key=openai_api_key)
                history_str = ", ".join([f"({e.event_type} on {e.target_component_id})" for e in all_interactions[-20:]])
                
                prompt = f"""
                You are an advanced Behavioral AI. Analyze the following sequence of user interactions:
                History: {history_str}
                
                Based on this pattern, predict the most likely next single action. 
                Return your response as a JSON object with 'predicted_event_type', 'predicted_target_component_id', and 'confidence' (float between 0 and 1).
                """
                
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[{"role": "system", "content": "You are a predictive JSON API."}, {"role": "user", "content": prompt}],
                    response_format={ "type": "json_object" }
                )
                
                prediction_data = json.loads(response.choices[0].message.content)
                return {
                    "predicted_next_event_type": prediction_data.get("predicted_event_type", "UNKNOWN"),
                    "predicted_target_component_id": prediction_data.get("predicted_target_component_id"),
                    "confidence": float(prediction_data.get("confidence", 0.5)),
                    "message": "AI Deep-Learning Prediction based on recent complex behavior sequences."
                }
            except Exception as e:
                print(f"[AI WARNING] LLM prediction failed, falling back to heuristic cluster: {e}")

        if len(all_interactions) < 2:
            return {"message": "Not enough interaction data to generate a prediction.", "confidence": 0.0}

        next_action_counter = Counter()
        total_occurrences = 0

        # Iterate through the sequence of events to find patterns
        for i in range(len(all_interactions) - 1):
            current_event = all_interactions[i]
            
            # Check if the current event in the history matches the user's last action
            if current_event.event_type == current_event_type and current_event.target_component_id == current_target_id:
                total_occurrences += 1
                next_event = all_interactions[i+1]
                action_key = (next_event.event_type, next_event.target_component_id)
                next_action_counter[action_key] += 1

        if not next_action_counter:
            return {"message": f"No historical patterns found following the event '{current_event_type}' on '{current_target_id}'.", "confidence": 0.0}

        most_common_action, count = next_action_counter.most_common(1)[0]
        predicted_event_type, predicted_target_id = most_common_action
        confidence = count / total_occurrences if total_occurrences > 0 else 0.0

        return {
            "predicted_next_event_type": predicted_event_type,
            "predicted_target_component_id": predicted_target_id,
            "confidence": confidence,
            "message": f"Based on {total_occurrences} occurrences in recent history, the predicted next action has a {confidence:.0%} confidence."
        }

    def generate_conversational_insight(self, db: Session, user_id: str, query: str) -> dict:
        """
        Generates a natural language response to a user's query based on their recent activity.
        This is a simple implementation of Conversational AI.
        """
        # Fetch recent interactions to provide context
        recent_interactions = db.query(models.UserInteractionEvent).filter(
            models.UserInteractionEvent.user_id == user_id
        ).order_by(desc(models.UserInteractionEvent.timestamp)).limit(10).all()

        if not recent_interactions:
            return {"answer": "I don't have any recent activity for you to answer questions about."}

        query_lower = query.lower()
        
        # --- Rule-based intent matching ---

        # 1. Handle date-based queries first for specificity
        if "yesterday" in query_lower:
            today = datetime.datetime.utcnow().date()
            yesterday = today - datetime.timedelta(days=1)
            start_of_yesterday = datetime.datetime.combine(yesterday, datetime.time.min).isoformat()
            end_of_yesterday = datetime.datetime.combine(yesterday, datetime.time.max).isoformat()

            yesterdays_interactions = db.query(models.UserInteractionEvent).filter(
                models.UserInteractionEvent.user_id == user_id,
                models.UserInteractionEvent.timestamp >= start_of_yesterday,
                models.UserInteractionEvent.timestamp <= end_of_yesterday
            ).all()

            if not yesterdays_interactions:
                return {"answer": "I found no recorded activity for you yesterday."}

            total_actions = len(yesterdays_interactions)
            type_counts = Counter(event.event_type for event in yesterdays_interactions)
            summary_str = ", ".join([f"'{event_type}' ({count} times)" for event_type, count in type_counts.items()])
            answer = f"Yesterday, you performed {total_actions} actions, including: {summary_str}."
            return {"answer": answer, "context": {"date_range": {"start": start_of_yesterday, "end": end_of_yesterday}, "action_count": total_actions}}

        elif "today" in query_lower:
            start_of_today = datetime.datetime.combine(datetime.datetime.utcnow().date(), datetime.time.min).isoformat()
            todays_interactions = db.query(models.UserInteractionEvent).filter(models.UserInteractionEvent.user_id == user_id, models.UserInteractionEvent.timestamp >= start_of_today).all()

            if not todays_interactions:
                return {"answer": "I haven't seen any activity from you yet today."}

            total_actions = len(todays_interactions)
            type_counts = Counter(event.event_type for event in todays_interactions)
            summary_str = ", ".join([f"'{event_type}' ({count} times)" for event_type, count in type_counts.items()])
            answer = f"So far today, you have performed {total_actions} actions, including: {summary_str}."
            return {"answer": answer, "context": {"date_range": {"start": start_of_today, "end": "now"}, "action_count": total_actions}}

        elif "last" in query_lower and ("days" in query_lower or "week" in query_lower):
            num_days = 0
            if "week" in query_lower:
                num_days = 7
            else:
                match = re.search(r'last\s+(\d+)\s+days', query_lower)
                if match:
                    num_days = int(match.group(1))

            if num_days > 0:
                today = datetime.datetime.utcnow().date()
                start_date = today - datetime.timedelta(days=num_days - 1)
                
                start_of_period = datetime.datetime.combine(start_date, datetime.time.min).isoformat()
                end_of_period = datetime.datetime.combine(today, datetime.time.max).isoformat()

                interactions = db.query(models.UserInteractionEvent).filter(
                    models.UserInteractionEvent.user_id == user_id,
                    models.UserInteractionEvent.timestamp >= start_of_period,
                    models.UserInteractionEvent.timestamp <= end_of_period
                ).all()

                if not interactions:
                    return {"answer": f"I found no recorded activity for you in the last {num_days} days."}

                total_actions = len(interactions)
                type_counts = Counter(event.event_type for event in interactions)
                summary_str = ", ".join([f"'{event_type}' ({count} times)" for event_type, count in type_counts.items()])
                day_str = "week" if num_days == 7 and "week" in query_lower else f"{num_days} days"
                answer = f"In the last {day_str}, you performed {total_actions} actions, including: {summary_str}."
                return {"answer": answer, "context": {"date_range": {"start": start_of_period, "end": end_of_period}, "action_count": total_actions}}

        # 2. Handle pattern/sequence analysis queries
        elif "sequence" in query_lower or "pattern" in query_lower or "common action" in query_lower:
            # Analyze a larger set of recent history for patterns
            all_interactions = list(reversed(db.query(models.UserInteractionEvent).filter(
                models.UserInteractionEvent.user_id == user_id
            ).order_by(desc(models.UserInteractionEvent.timestamp)).limit(500).all()))

            if len(all_interactions) < 2:
                return {"answer": "I don't have enough interaction data to find any action patterns yet."}

            sequence_counter = Counter()
            for i in range(len(all_interactions) - 1):
                event1 = all_interactions[i]
                event2 = all_interactions[i+1]
                
                # Create a readable key for the sequence
                action1_desc = f"a {event1.event_type} on '{event1.target_component_id or 'an unspecified target'}'"
                action2_desc = f"a {event2.event_type} on '{event2.target_component_id or 'an unspecified target'}'"
                sequence_key = (action1_desc, action2_desc)
                
                sequence_counter[sequence_key] += 1

            if not sequence_counter:
                return {"answer": "I couldn't find any repeating action sequences in your recent history."}

            most_common_sequence, count = sequence_counter.most_common(1)[0]
            action1, action2 = most_common_sequence
            
            answer = f"Your most frequent action sequence is: performing {action1}, followed immediately by {action2}. I've seen this pattern {count} times in your recent history."
            return {"answer": answer, "context": {"sequence": most_common_sequence, "count": count, "analyzed_interactions": len(all_interactions)}}

        # 3. Handle other query types
        elif "what did i do last" in query_lower or "last action" in query_lower:
            last_event = recent_interactions[0]
            answer = f"Your last action was a '{last_event.event_type}'"
            if last_event.target_component_id:
                answer += f" on the component '{last_event.target_component_id}'."
            else:
                answer += "."
            return {"answer": answer, "context": {"event_id": last_event.event_id}}

        elif "how many times" in query_lower:
            if "click" in query_lower:
                event_type_to_count = "BUTTON_CLICK"
            elif "view" in query_lower:
                event_type_to_count = "SCREEN_VIEW"
            else:
                return {"answer": "I can count event types like 'clicks' or 'views'. Please be more specific."}
            
            count = sum(1 for event in recent_interactions if event.event_type == event_type_to_count)
            return {"answer": f"In your recent activity, you performed the action '{event_type_to_count}' {count} times."}

        elif "predict" in query_lower or "what's next" in query_lower:
            last_event = recent_interactions[0]
            prediction = self.generate_predictive_insight(db, user_id, last_event.event_type, last_event.target_component_id)
            
            if prediction.get("predicted_next_event_type"):
                confidence = prediction.get('confidence', 0.0)
                answer = f"Based on your habits, I predict with {confidence:.0%} confidence that your next action will be a '{prediction['predicted_next_event_type']}' on '{prediction.get('predicted_target_component_id', 'an unspecified target')}'."
                return {"answer": answer, "context": prediction}
            else:
                return {"answer": "I couldn't predict your next action based on your recent activity.", "context": prediction}

        else:
            return {"answer": "I can answer questions about your recent activity, like 'What did I do last?' or 'Predict my next action'."}

    def get_interaction_statistics(self, db: Session) -> dict:
        """
        Retrieves system-wide statistics on user interactions.
        """
        total_interactions = db.query(models.UserInteractionEvent).count()
        
        total_unique_users = db.query(func.count(distinct(models.UserInteractionEvent.user_id))).scalar()

        stats_query = db.query(
            models.UserInteractionEvent.event_type,
            func.count(models.UserInteractionEvent.event_id).label('count')
        ).group_by(
            models.UserInteractionEvent.event_type
        ).order_by(
            desc('count')
        ).all()

        return {
            "total_interactions": total_interactions,
            "total_unique_users": total_unique_users or 0,
            "stats_by_event_type": stats_query
        }

    def summarize_interaction_stats_for_logging(self, db: Session, triggered_by: str) -> dict:
        """
        Retrieves system-wide statistics on user interactions and logs them as a maintenance task.
        Returns the statistics that were logged.
        """
        task_name = "summarize_interaction_stats"
        start_time = datetime.datetime.utcnow()
        try:
            stats = self.get_interaction_statistics(db)
            
            # The stats dict is already in a good format for the JSONB summary field.
            duration_ms = int((datetime.datetime.utcnow() - start_time).total_seconds() * 1000)
            self._log_task(db, task_name, "SUCCESS", triggered_by, summary=stats, duration_ms=duration_ms)
            
            return stats
        except Exception as e:
            db.rollback()
            duration_ms = int((datetime.datetime.utcnow() - start_time).total_seconds() * 1000)
            self._log_task(db, task_name, "FAILED", triggered_by, details=str(e), duration_ms=duration_ms)
            raise e

    def cleanup_old_interaction_events(self, db: Session, retention_days: int, triggered_by: str) -> int:
        """
        Finds and permanently deletes old user interaction events.
        This is a cleanup task, not an archival. Records are permanently deleted.

        Returns the number of events deleted.
        """
        task_name = "cleanup_interaction_events"
        start_time = datetime.datetime.utcnow()
        try:
            if retention_days < 30:
                raise ValueError("Retention period must be at least 30 days for safety.")

            cutoff_date = datetime.datetime.utcnow() - datetime.timedelta(days=retention_days)
            cutoff_date_str = cutoff_date.isoformat()

            events_to_delete_query = db.query(models.UserInteractionEvent).filter(
                models.UserInteractionEvent.timestamp < cutoff_date_str
            )
            deleted_count = events_to_delete_query.delete(synchronize_session=False)
            db.commit()

            summary = {"deleted_count": deleted_count, "retention_days": retention_days}
            duration_ms = int((datetime.datetime.utcnow() - start_time).total_seconds() * 1000)
            self._log_task(db, task_name, "SUCCESS", triggered_by, summary=summary, duration_ms=duration_ms)
            return deleted_count
        except Exception as e:
            db.rollback()
            duration_ms = int((datetime.datetime.utcnow() - start_time).total_seconds() * 1000)
            self._log_task(db, task_name, "FAILED", triggered_by, details=str(e), duration_ms=duration_ms)
            raise e

    def clear_user_interaction_history(self, db: Session, user_id: str, triggered_by: str) -> int:
        """
        Permanently deletes all interaction history for a specific user.
        This is a destructive operation for privacy/data deletion requests.

        Returns the number of events deleted.
        """
        task_name = "clear_user_interaction_history"
        start_time = datetime.datetime.utcnow()
        try:
            events_to_delete_query = db.query(models.UserInteractionEvent).filter(
                models.UserInteractionEvent.user_id == user_id
            )
            
            deleted_count = events_to_delete_query.delete(synchronize_session=False)
            db.commit()

            summary = {"deleted_count": deleted_count, "user_id": user_id}
            duration_ms = int((datetime.datetime.utcnow() - start_time).total_seconds() * 1000)
            self._log_task(db, task_name, "SUCCESS", triggered_by, summary=summary, duration_ms=duration_ms)
            return deleted_count
        except Exception as e:
            db.rollback()
            duration_ms = int((datetime.datetime.utcnow() - start_time).total_seconds() * 1000)
            self._log_task(db, task_name, "FAILED", triggered_by, details=str(e), summary={"user_id": user_id}, duration_ms=duration_ms)
            raise e

    def get_cleared_user_history(self, db: Session) -> List[Dict]:
        """
        Retrieves a list of all unique users who have had their interaction history cleared,
        showing the most recent clearance event for each user.
        """
        logs = db.query(models.MaintenanceTaskLog).filter(
            models.MaintenanceTaskLog.task_name == 'clear_user_interaction_history'
        ).order_by(desc(models.MaintenanceTaskLog.triggered_at)).all()

        cleared_users_map = {}
        for log in logs:
            # Ensure the log is successful and has the user_id in summary
            if log.status == "SUCCESS" and log.summary and "user_id" in log.summary:
                user_id = log.summary.get("user_id")
                if user_id and user_id not in cleared_users_map:
                    cleared_users_map[user_id] = {
                        "user_id": user_id,
                        "last_cleared_at": log.triggered_at,
                        "cleared_by": log.triggered_by
                    }
        
        return list(cleared_users_map.values())

    def generate_workflow_from_prompt(self, prompt: str, workflow_name: str, country_code: str) -> dict:
        """
        Parses a natural language prompt to generate a structured workflow manifest.
        This simulates the "NLP Prompt-to-Canvas" feature with jurisdictional awareness.
        """
        # A simple keyword-to-node mapping, with country-specific overrides.
        NODE_TEMPLATES = {
            "default": {
                "ingest": {"node_title": "Data Ingestion Gateway", "node_code": "INGEST", "rules_applied": ["VALIDATE_FORMAT"]},
                "upload": {"node_title": "File Upload Node", "node_code": "UPLOAD", "rules_applied": ["VALIDATE_FILENAME"]},
                "compliance": {"node_title": "Standard Compliance Check", "node_code": "COMPLIANCE", "rules_applied": ["GLOBAL_SANCTIONS_V1"]},
                "approve": {"node_title": "Manager Approval Screen", "node_code": "APPROVAL", "screen_template": "MANAGER_APPROVAL_FORM"},
                "enrich": {"node_title": "Data Enrichment Step", "node_code": "ENRICH", "api_triggers": ["GET_CUSTOMER_DATA"]},
                "ledger": {"node_title": "Post to General Ledger", "node_code": "POST_LEDGER", "api_triggers": ["CORE_BANKING_POST_API"]},
                "settle": {"node_title": "Final Settlement", "node_code": "SETTLE", "events_broadcast": ["WORKFLOW_COMPLETED"]},
            },
            "DE": { # Germany-specific overrides
                "compliance": {"node_title": "BaFin Compliance & AML Check", "node_code": "COMPLIANCE_DE", "rules_applied": ["BAFIN_AML_2024", "GLOBAL_SANCTIONS_V1"]},
            },
            "US": { # US-specific overrides
                "compliance": {"node_title": "US KYC & AML Check", "node_code": "COMPLIANCE_US", "rules_applied": ["US_PATRIOT_ACT_V4", "GLOBAL_SANCTIONS_V1"]},
            }
        }

        # Merge default templates with country-specific ones
        country_templates = NODE_TEMPLATES.get(country_code.upper(), {})
        final_templates = {**NODE_TEMPLATES["default"], **country_templates}

        prompt_lower = prompt.lower()
        detected_nodes = []
        
        # Simple keyword detection to determine which nodes to create
        for keyword, template in final_templates.items():
            if keyword in prompt_lower:
                detected_nodes.append(template)

        if not detected_nodes:
            raise ValueError("Could not generate a workflow. The prompt did not contain any recognizable keywords like 'ingest', 'compliance', 'approve', or 'ledger'.")

        # Build the nodes and edges for the manifest
        nodes = []
        edges = []
        last_node_id = None
        
        for i, node_template in enumerate(detected_nodes):
            node_id = f"NODE-{uuid.uuid4().hex[:8].upper()}"
            node = {
                "sequence_number": i + 1,
                "node_title": node_template["node_title"],
                "node_code": node_template["node_code"],
                "canvas_x_position": 50 + (i * 250), # Arrange nodes horizontally
                "canvas_y_position": 150,
                "rules_applied": node_template.get("rules_applied"),
                "api_triggers": node_template.get("api_triggers"),
                "events_broadcast": node_template.get("events_broadcast"),
                "screen_template": node_template.get("screen_template"),
            }
            nodes.append(node)

            if last_node_id:
                edge = {
                    "source_node_id": last_node_id,
                    "target_node_id": node_id,
                }
                edges.append(edge)
            
            last_node_id = node_id

        manifest = {"workflow_name": workflow_name, "domain_scope": "AI_GENERATED", "product_context": "AI Prompt-to-Canvas", "nodes": nodes, "edges": edges}

        return {"message": f"Successfully generated a {len(nodes)}-step workflow manifest from the prompt.", "generated_manifest": manifest}

    def generate_rule_from_prompt(self, db: Session, prompt: str, current_user: schemas.CurrentUser) -> dict:
        """
        Parses a natural language prompt to generate a complete, reusable BusinessRuleSet.
        This simulates the core of the "Infinity AI Assistant" for rule generation.
        """
        # 1. NLU: Use regex to parse the "IF...THEN..." structure.
        # Example: "If Payment Amount is greater than 2000000 then a director needs to approve"
        match = re.search(r"if\s+(.+?)\s+(is greater than|is less than|is equal to|is not equal to)\s+([\d\.]+)\s+then\s+(.+)", prompt, re.IGNORECASE)
        if not match:
            raise ValueError("Could not parse the prompt. Please use the format 'IF <field name> <operator> <value> THEN <action>'.")

        field_name_str, operator_str, value_str, action_str = match.groups()
        value = float(value_str)

        # 2. Entity Resolution: Find the field in the registry.
        field_search_term = f"%{field_name_str.replace(' ', '%')}%"
        target_field = db.query(models.ISOFieldDefinition).filter(
            models.ISOFieldDefinition.client_business_name.ilike(field_search_term)
        ).first()
        if not target_field:
            raise ValueError(f"Could not find a field matching '{field_name_str}' in the ISO Field Registry.")
        
        technical_field_name = target_field.technical_sys_name

        # 3. Component Generation: Create the BusinessRuleSet payload.
        operator_map = {
            "is greater than": "GREATER_THAN",
            "is less than": "LESS_THAN",
            "is equal to": "EQUAL_TO",
            "is not equal to": "NOT_EQUAL_TO",
        }
        operator = operator_map.get(operator_str.lower())
        if not operator:
            raise ValueError(f"Unsupported operator: '{operator_str}'")

        # Construct the rule definition using our enhanced schemas
        rule_definition = schemas.BusinessRuleSet(
            business_name=f"AI: {field_name_str} check",
            token_code=f"BRE-AI-{uuid.uuid4().hex[:6].upper()}",
            description=f"Generated by AI from prompt: '{prompt}'",
            rules=[
                schemas.BusinessRule(
                    priority=100,
                    conditions=[
                        schemas.RuleCondition(
                            left_hand_side=schemas.RuleConditionOperand(source_fields=[technical_field_name]),
                            operator=operator,
                            right_hand_side=schemas.RuleConditionOperand(static_value=value)
                        )
                    ],
                    actions=[] # The action is handled by the orchestration step
                )
            ]
        )

        # 4. Save the new BusinessRuleSet to the database
        new_rule_set_db = models.BusinessRuleSet(
            rule_set_id=f"BRE-{uuid.uuid4().hex[:8].upper()}",
            business_name=rule_definition.business_name,
            token_code=rule_definition.token_code,
            description=rule_definition.description,
            definition=rule_definition.dict(),
            created_at=datetime.datetime.utcnow().isoformat(),
            created_by=current_user.id
        )
        db.add(new_rule_set_db)
        db.commit()

        # 5. Suggest the Workflow Node to the user
        # The AI interprets the action part of the prompt to suggest a new node.
        suggested_node = None
        suggested_edge_condition = None
        notes = []
        action_lower = action_str.lower()
        if "approve" in action_lower or "approval" in action_lower:
            # A more advanced AI could look up roles like 'director'
            # For now, we create a generic, human-in-the-loop approval node.
            suggested_node = schemas.WorkflowNodeCreate(
                sequence_number=99, # Suggest a high number to place it later in a flow
                node_title=f"AI Suggested: {action_str.strip().title()}",
                node_code="HUMAN_APPROVAL",
                screen_template="GENERIC_APPROVAL_FORM_V1", # A standard screen for approvals
                orchestration_steps=[] # This node's action is human, so no further orchestration
            )
            # Also suggest the exact JSON for the conditional edge
            suggested_edge_condition = {
                "type": "BUSINESS_RULE",
                "token_code": rule_definition.token_code
            }

            notes = [
                f"The action '{action_str}' was interpreted as needing a new human approval node.",
                "You should add this suggested node to your workflow canvas.",
                f"Then, create a conditional edge from your previous node to this new node. Use the generated rule '{rule_definition.token_code}' as the edge's condition."
            ]
        else:
            notes = [
                f"The condition part of your prompt was understood and rule '{rule_definition.token_code}' was created.",
                f"The action '{action_str}' was not automatically converted into a node. Please configure the 'THEN' part of your logic manually."
            ]

        return {
            "message": "Successfully generated a new Business Rule Set from your prompt.",
            "generated_rule_token": rule_definition.token_code,
            "suggested_workflow_node": suggested_node,
            "suggested_edge_condition": suggested_edge_condition,
            "notes": notes
        }

    def parse_and_execute_command(self, db: Session, prompt: str, current_user: schemas.CurrentUser) -> dict:
        """
        Parses a natural language command and executes the corresponding action.
        This is the central "brain" of the AI Assistant.
        """
        prompt_lower = prompt.lower()

        # --- Intent Routing ---
        # A real implementation would use a more sophisticated NLU model.
        # For now, we use regex and keyword matching.

        # Intent: Add a new currency
        currency_match = re.search(r"add\s+([a-zA-Z]{3})\s+currency", prompt_lower)
        if currency_match:
            currency_code = currency_match.group(1).upper()
            
            # Check if currency already exists
            existing = db.query(models.CurrencyMaster).filter(models.CurrencyMaster.currency_code == currency_code).first()
            if existing:
                raise ValueError(f"Cannot add currency. '{currency_code}' already exists in the Currency Master.")

            # Create the new currency object
            # In a real system, the AI might ask for more details. Here we use defaults.
            new_currency = models.CurrencyMaster(
                currency_code=currency_code,
                currency_name=f"{currency_code} (AI Added)",
                fraction_digits=2,
                source_currency_code="USD",
                target_currency_code="USD",
                exchange_rate=1.0,
                created_at=datetime.datetime.utcnow().isoformat(),
                created_by=f"AI_ASSISTANT_FOR_{current_user.id}"
            )
            db.add(new_currency)
            db.commit()
            db.refresh(new_currency)

            message = f"Successfully added new currency '{currency_code}' to the Currency Master."
            details = {"currency_code": new_currency.currency_code}

            # Handle secondary intent: "...and send for approval"
            if "approve" in prompt_lower or "approval" in prompt_lower:
                event_payload = {
                    "entity_type": "CurrencyMaster",
                    "entity_id": new_currency.currency_code,
                    "action_required": "APPROVAL",
                    "initiated_by": current_user.id
                }
                asyncio.run(global_event_bus.broadcast(SystemEvent(
                    event_type="MASTERS_ENTRY_PENDING_APPROVAL",
                    source_context="AI_Assistant",
                    payload=event_payload
                )))
                message += " A notification has been sent to the checker for approval."

            return {
                "status": "SUCCESS",
                "message": message,
                "executed_action": "CREATE_CURRENCY",
                "details": details
            }

        # If no intent is matched
        raise ValueError("I'm sorry, I didn't understand that command. I can currently add currencies (e.g., 'add JPY currency').")

    def run_and_log_behavioral_profile_update(self, db: Session, triggered_by: str) -> dict:
        """
        A wrapper function for the scheduler that executes the profile update and logs the result.
        Upgraded to use Asynchronous Checkpointing to prevent OOM and ensure Resumability (Pillar 8).
        """
        from itertools import groupby
        
        task_name = "update_behavioral_profiles"
        start_time = datetime.datetime.utcnow()
        
        # 1. Resume Check: Look for an interrupted job
        job = db.query(models.BehavioralProfileUpdateJob).filter(
            models.BehavioralProfileUpdateJob.status.in_(["PENDING", "PROCESSING"])
        ).first()

        if not job:
            job_id = f"BEHAV-JOB-{uuid.uuid4().hex[:8].upper()}"
            job = models.BehavioralProfileUpdateJob(
                job_id=job_id,
                status="PROCESSING",
                created_at=start_time.isoformat()
            )
            db.add(job)
            db.commit()

        try:
            # 2. Establish Bounds: Get all unique users
            unique_users = db.query(models.UserInteractionEvent.user_id).distinct().all()
            user_ids = [u[0] for u in unique_users]
            
            if job.total_users is None:
                job.total_users = len(user_ids)
                db.commit()

            CHUNK_SIZE = 100 # Process 100 users per checkpoint to prevent OOM
            start_index = job.processed_users or 0
            updated_profiles_count = 0

            # 3. Process Chunk Vectorially and Checkpoint
            for i in range(start_index, len(user_ids), CHUNK_SIZE):
                chunk_user_ids = user_ids[i:i + CHUNK_SIZE]
                
                # Fetch interactions ONLY for this chunk of users (Preventing OOM)
                interactions = db.query(models.UserInteractionEvent).filter(
                    models.UserInteractionEvent.user_id.in_(chunk_user_ids)
                ).order_by(models.UserInteractionEvent.user_id, models.UserInteractionEvent.timestamp).all()

                for user_id, interactions_group in groupby(interactions, key=lambda x: x.user_id):
                    user_interactions = list(interactions_group)
                    
                    # --- Calculate Ranked Journeys ---
                    journey_counts = Counter(e.target_component_id for e in user_interactions if e.target_component_id)
                    ranked_journeys = [{"journey_id": j, "interaction_count": c} for j, c in journey_counts.most_common(5)]

                    # --- Calculate Common Devices ---
                    device_fingerprints = [e.payload.get("device_fingerprint") for e in user_interactions if e.payload and e.payload.get("device_fingerprint")]
                    common_devices = [{"fingerprint": d, "count": c} for d, c in Counter(device_fingerprints).most_common(3)]

                    # --- Calculate Typical Locations ---
                    locations = [e.payload.get("geo_location") for e in user_interactions if e.payload and e.payload.get("geo_location")]
                    typical_locations = [{"location": loc, "count": c} for loc, c in Counter(locations).most_common(3)]

                    # --- Create or Update the Profile ---
                    profile_data = {
                        "user_id": user_id,
                        "ranked_journeys": ranked_journeys,
                        "common_devices": common_devices,
                        "typical_locations": typical_locations,
                        "last_calculated_at": datetime.datetime.utcnow().isoformat(),
                    }
                    
                    new_profile = models.CustomerBehavioralProfile(**profile_data)
                    db.merge(new_profile)
                    updated_profiles_count += 1
                
                # 4. Save State Checkpoint! If server crashes after this, we resume at next chunk.
                job.processed_users = i + len(chunk_user_ids)
                db.commit()

            # 5. Finalize Job
            job.status = "COMPLETED"
            job.completed_at = datetime.datetime.utcnow().isoformat()
            db.commit()

            summary = {"profiles_updated": updated_profiles_count, "total_users_analyzed": len(user_ids), "job_id": job.job_id}
            duration_ms = int((datetime.datetime.utcnow() - start_time).total_seconds() * 1000)
            self._log_task(db, task_name, "SUCCESS", triggered_by, summary=summary, duration_ms=duration_ms)
            return summary

        except Exception as e:
            db.rollback()
            try:
                job.status = "FAILED"
                job.error_message = str(e)
                db.commit()
            except Exception:
                pass
            duration_ms = int((datetime.datetime.utcnow() - start_time).total_seconds() * 1000)
            self._log_task(db, task_name, "FAILED", triggered_by, details=str(e), duration_ms=duration_ms)
            raise e

    def generate_insight_from_prompt(self, db: Session, prompt: str, current_user: schemas.CurrentUser) -> dict:
        """
        Parses a natural language prompt to generate a complete InsightDefinition blueprint.
        This is a placeholder for a more advanced NLU model.
        """
        prompt_lower = prompt.lower()
        
        # --- Simple Intent Matching for "Similar Subscriptions" example ---
        if "similar" in prompt_lower and "subscription" in prompt_lower:
            insight_name = "AI: Similar Subscription Detector"
            insight_code = f"INSIGHT-AI-{uuid.uuid4().hex[:4].upper()}"

            # This insight would be triggered by a new transaction event
            trigger_type = "EVENT"
            trigger_config = {"event_type": "NEW_TRANSACTION"}

            # The analysis would involve a business rule
            # This rule would need to be created separately, but the AI can define the step
            analysis_steps = [
                schemas.OrchestrationStep(
                    sequence_number=10,
                    step_type="BUSINESS_RULE",
                    target_token="BRE-CHECK-DUPLICATE-SUBSCRIPTIONS-V1" # A pre-existing rule
                )
            ]

            insight_payload = schemas.InsightDefinitionCreate(
                insight_name=insight_name,
                insight_code=insight_code,
                description=f"Generated by AI from prompt: '{prompt}'",
                trigger_type=trigger_type,
                trigger_config=trigger_config,
                analysis_steps=analysis_steps
            )

            return {"message": "Successfully generated a blueprint for the 'Similar Subscriptions' insight.", "generated_insight_blueprint": insight_payload.dict()}

        raise ValueError("I can currently only generate a blueprint for the 'similar subscriptions' insight.")

    def generate_screen_from_wireframe(self, db: Session, image_base64: str, mime_type: str) -> dict:
        """
        Uses Vision-capable LLMs to extract UI components from a wireframe image
        and attempts to auto-map them to the ISO Field Registry.
        """
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if not openai_api_key:
            raise ValueError("OpenAI API key not configured. Cannot perform wireframe extraction.")

        # Handle Document Conversions (PDF to Image)
        if mime_type == 'application/pdf':
            try:
                pdf_bytes = base64.b64decode(image_base64)
                doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                if len(doc) > 0:
                    page = doc.load_page(0) # Render the first page of the PDF
                    pix = page.get_pixmap(dpi=150)
                    image_bytes = pix.tobytes("jpeg")
                    image_base64 = base64.b64encode(image_bytes).decode('utf-8')
                    mime_type = 'image/jpeg'
                else:
                    raise ValueError("Uploaded PDF has no pages.")
                doc.close()
            except Exception as e:
                raise ValueError(f"Failed to parse PDF file into image: {str(e)}")

        # Check if structured data (CSV/Excel) to extract headers and samples
        text_payload = None
        if mime_type in ['text/csv', 'application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'] or mime_type.endswith('csv') or mime_type.endswith('excel'):
            try:
                file_bytes = base64.b64decode(image_base64)
                if 'csv' in mime_type:
                    decoded_file = file_bytes.decode('utf-8')
                    csv_reader = csv.reader(io.StringIO(decoded_file))
                    rows = [row for _, row in zip(range(5), csv_reader)]
                    text_payload = "Data Sample:\n" + "\n".join([", ".join(r) for r in rows if r])
                else:
                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                    sheet = wb.active
                    rows = [row for _, row in zip(range(5), sheet.iter_rows(values_only=True))]
                    text_payload = "Data Sample:\n" + "\n".join([", ".join([str(c) for c in r if c is not None]) for r in rows if r])
            except Exception as e:
                raise ValueError(f"Failed to parse structured file: {str(e)}")

        # 1. Fetch available ISO fields to provide exact mapping context to the LLM
        fields = db.query(models.ISOFieldDefinition).limit(200).all()
        field_context = ", ".join([f"'{f.technical_sys_name}' (business name: '{f.client_business_name}')" for f in fields])

        try:
            client = OpenAI(api_key=openai_api_key)
            prompt = f"""
            You are an expert UX Developer and Business Architect. Analyze the provided wireframe image OR data sample.
            If an image is provided, identify all form fields, inputs, dropdowns, date pickers, and labels.
            If a data sample is provided, infer the logical form fields required to collect or display this data.
            For each field, attempt to map it to the closest fitting ISO Field from this available registry list: [{field_context}].
            If there is no logical match, leave 'field_binding' as an empty string "".
            
            Return your response STRICTLY as a JSON object matching this schema:
            {{
                "components": [
                    {{
                        "component_type": "text_input" | "number_input" | "dropdown" | "date_picker" | "label",
                        "label_token": "The display label or title found in the image",
                        "field_binding": "The exact technical_sys_name from the registry, or empty string",
                        "category": "USER_DEFINED",
                        "requirement_status": "MANDATORY" | "NON_MANDATORY"
                    }}
                ]
            }}
            """
            
            if text_payload:
                messages = [
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": f"Analyze this data sample and build a UI screen to input or display this data:\n{text_payload}"}
                ]
            else:
                messages = [{"role": "user", "content": [{"type": "text", "text": prompt}, {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_base64}"}}]}]
            
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=messages,
                response_format={ "type": "json_object" }
            )
            
            return json.loads(response.choices[0].message.content)
        except Exception as e:
            raise ValueError(f"AI Vision extraction failed: {str(e)}")

    def generate_field_translations(self, business_name: str, domain_category: Optional[str] = "General Banking") -> dict:
        """
        Acts as a Global Multilingual Dictionary. Translates a given English business term
        into standard international locales using precise financial terminology.
        """
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if not openai_api_key:
            raise ValueError("OpenAI API key not configured. Cannot perform translation.")

        try:
            client = OpenAI(api_key=openai_api_key)
            prompt = f"""
            You are an expert financial translator and localization engineer. 
            Translate the following English banking field name into highly accurate, professional financial terminology for the requested locales.
            
            Field Name: "{business_name}"
            Business Context: "{domain_category}"
            
            Return your response STRICTLY as a JSON object where keys are the locale codes ("es", "fr", "de", "zh", "ja", "ar") and the values are the translated strings.
            """
            
            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                response_format={ "type": "json_object" }
            )
            
            return json.loads(response.choices[0].message.content)
        except Exception as e:
            raise ValueError(f"AI translation generation failed: {str(e)}")

    def generate_report_from_prompt(self, db: Session, prompt: str, current_user: schemas.CurrentUser) -> dict:
        """
        Parses a natural language prompt to generate a complete ReportBlueprint.
        """
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if not openai_api_key:
            raise ValueError("OpenAI API key not configured. Cannot perform AI report generation.")

        # Provide the LLM with the context of all available fields from the bloodsteam
        fields = db.query(models.ISOFieldDefinition).limit(200).all()
        field_context = ", ".join([f"'{f.technical_sys_name}' (business name: '{f.client_business_name}')" for f in fields])
        
        try:
            client = OpenAI(api_key=openai_api_key)
            system_prompt = f"""
            You are an expert Data Analyst and BI Dashboard Designer. Analyze the provided user prompt.
            Determine the required charts/widgets and map them to the closest fitting ISO Field from this available registry list: [{field_context}].
            
            Return your response STRICTLY as a JSON object matching this schema:
            {{
                "report_name": "A suitable name for the dashboard",
                "description": "A description of the dashboard",
                "widgets": [
                    {{
                        "widget_id": "WGT-12345",
                        "chart_type": "PIE_CHART | BAR_CHART | LINE_CHART | DATA_GRID | KPI_CARD",
                        "title": "Widget title",
                        "data_source_entity": "EvidencePacketRegistry",
                        "x_axis_field": "technical_sys_name for grouping/X-axis",
                        "y_axis_field": "technical_sys_name for measurement/Y-axis",
                        "aggregation_method": "SUM | COUNT | AVG",
                        "grid_layout": {{"x": 0, "y": 0, "w": 6, "h": 4}}
                    }}
                ]
            }}
            """
            
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": prompt}],
                response_format={ "type": "json_object" }
            )
            
            parsed_response = json.loads(response.choices[0].message.content)
            
            # Construct the Pydantic-compliant ReportBlueprint payload
            report_blueprint = schemas.ReportBlueprintCreate(
                report_name=parsed_response.get("report_name", "AI Generated Dashboard"),
                description=parsed_response.get("description", f"Generated by AI from prompt: '{prompt}'"),
                is_third_party_embedded=False,
                expose_as_headless_api=False,
                widgets=[schemas.ReportWidgetConfig(**w) for w in parsed_response.get("widgets", [])]
            )

            return {
                "message": "Successfully generated a blueprint for the reporting dashboard.",
                "generated_report_blueprint": report_blueprint.dict(),
                "notes": ["You can save this blueprint and edit it further in the Report Designer Canva."]
            }
        except Exception as e:
            raise ValueError(f"AI report generation failed: {str(e)}")

    def generate_report_from_image(self, db: Session, image_base64: str, mime_type: str, current_user: schemas.CurrentUser) -> dict:
        """
        Uses Vision-capable LLMs to extract charts and layout from a report screenshot/image
        and map them to the ISO Field Registry to generate a ReportBlueprint.
        """
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if not openai_api_key:
            raise ValueError("OpenAI API key not configured. Cannot perform wireframe extraction.")

        # Handle Document Conversions (PDF to Image)
        if mime_type == 'application/pdf':
            try:
                pdf_bytes = base64.b64decode(image_base64)
                doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                if len(doc) > 0:
                    page = doc.load_page(0) # Render the first page of the PDF
                    pix = page.get_pixmap(dpi=150)
                    image_bytes = pix.tobytes("jpeg")
                    image_base64 = base64.b64encode(image_bytes).decode('utf-8')
                    mime_type = 'image/jpeg'
                else:
                    raise ValueError("Uploaded PDF has no pages.")
                doc.close()
            except Exception as e:
                raise ValueError(f"Failed to parse PDF file into image: {str(e)}")

        # Check if structured data (CSV/Excel) to extract headers and samples
        text_payload = None
        if mime_type in ['text/csv', 'application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'] or mime_type.endswith('csv') or mime_type.endswith('excel'):
            try:
                file_bytes = base64.b64decode(image_base64)
                if 'csv' in mime_type:
                    decoded_file = file_bytes.decode('utf-8')
                    csv_reader = csv.reader(io.StringIO(decoded_file))
                    rows = [row for _, row in zip(range(5), csv_reader)]
                    text_payload = "Data Sample:\n" + "\n".join([", ".join(r) for r in rows if r])
                else:
                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                    sheet = wb.active
                    rows = [row for _, row in zip(range(5), sheet.iter_rows(values_only=True))]
                    text_payload = "Data Sample:\n" + "\n".join([", ".join([str(c) for c in r if c is not None]) for r in rows if r])
            except Exception as e:
                raise ValueError(f"Failed to parse structured file: {str(e)}")

        fields = db.query(models.ISOFieldDefinition).limit(200).all()
        field_context = ", ".join([f"'{f.technical_sys_name}' (business name: '{f.client_business_name}')" for f in fields])

        try:
            client = OpenAI(api_key=openai_api_key)
            prompt = f"""
            You are an expert BI Dashboard Designer. Analyze the provided image of a report OR the structured data sample.
            If an image is provided, identify the charts, data grids, and KPI cards present.
            If a data sample is provided, deduce the most insightful charts and aggregations to visualize this data.
            For each widget, determine the best fitting ISO Field from this registry for the X and Y axes: [{field_context}].
            Estimate a reasonable React-Grid-Layout coordinate object `grid_layout` (x, y, w, h) based on its visual position (assume a 12-column grid format).
            
            Return your response STRICTLY as a JSON object matching this schema:
            {{
                "report_name": "A suitable name based on the image",
                "description": "Description of the visual report",
                "widgets": [
                    {{
                        "widget_id": "WGT-123456",
                        "chart_type": "PIE_CHART" | "BAR_CHART" | "LINE_CHART" | "DATA_GRID" | "KPI_CARD",
                        "title": "Widget title extracted from image",
                        "data_source_entity": "EvidencePacketRegistry",
                        "x_axis_field": "technical_sys_name",
                        "y_axis_field": "technical_sys_name",
                        "aggregation_method": "SUM" | "COUNT" | "AVG",
                        "grid_layout": {{"x": 0, "y": 0, "w": 6, "h": 4}}
                    }}
                ]
            }}
            """
            
            if text_payload:
                messages = [
                    {"role": "system", "content": prompt},
                    {"role": "user", "content": f"Analyze this dataset and build a business intelligence dashboard with appropriate charts:\n{text_payload}"}
                ]
            else:
                messages = [{"role": "user", "content": [{"type": "text", "text": prompt}, {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_base64}"}}]}]
            
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=messages,
                response_format={ "type": "json_object" }
            )
            
            parsed_response = json.loads(response.choices[0].message.content)
            
            report_blueprint = schemas.ReportBlueprintCreate(
                report_name=parsed_response.get("report_name", "AI Extracted Dashboard"),
                description=parsed_response.get("description", "Extracted from uploaded report image."),
                is_third_party_embedded=False,
                expose_as_headless_api=False,
                widgets=[schemas.ReportWidgetConfig(**w) for w in parsed_response.get("widgets", [])]
            )
            
            return {
                "message": "Successfully extracted report layout and bindings.",
                "generated_report_blueprint": report_blueprint.dict()
            }
        except Exception as e:
            raise ValueError(f"AI Vision extraction failed: {str(e)}")

    def auto_map_file(self, db: Session, file_bytes: bytes, filename: str) -> dict:
        """
        Analyzes an uploaded file (CSV/Excel), extracts headers and sample data,
        and uses an LLM to auto-map them to the ISO Field Registry.
        """
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if not openai_api_key:
            raise ValueError("OpenAI API key not configured. Cannot perform auto-mapping.")

        headers = []
        sample_row = []
        file_type = "CSV"
        
        try:
            if filename.endswith('.csv'):
                decoded_file = file_bytes.decode('utf-8')
                csv_reader = csv.reader(io.StringIO(decoded_file))
                headers = next(csv_reader, [])
                sample_row = next(csv_reader, [])
                file_type = "CSV"
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                sheet = wb.active
                rows = list(sheet.iter_rows(values_only=True, max_row=2))
                if len(rows) > 0: headers = [str(h) if h is not None else f"Column_{i}" for i, h in enumerate(rows[0])]
                if len(rows) > 1: sample_row = [str(c) if c is not None else "" for c in rows[1]]
                file_type = "XLSX"
            else:
                raise ValueError("Unsupported file format for auto-mapping. Use CSV or Excel.")
        except Exception as e:
            raise ValueError(f"Failed to parse file for auto-mapping: {str(e)}")

        fields = db.query(models.ISOFieldDefinition).limit(500).all()
        field_context = ", ".join([f"'{f.technical_sys_name}' (business name: '{f.client_business_name}')" for f in fields])

        try:
            client = OpenAI(api_key=openai_api_key)
            prompt = f"""
            You are an expert Data Architect. I am providing you with the headers and a sample row from a {file_type} file.
            Map each header to the most appropriate ISO Field from this registry: [{field_context}].
            If no field closely matches the header's intent or data type, set 'is_new_field_required' to true, 'suggested_iso_field' to null, and infer the 'inferred_data_type' (e.g., 'Text', 'Amount', 'Date', 'Decimal', 'Alphanumeric').
            
            Headers: {headers}
            Sample Row: {sample_row}
            
            Return STRICTLY a JSON object matching this schema:
            {{ "suggested_mappings": [ {{ "source_path": "header_name", "suggested_iso_field": "technical_sys_name_or_null", "confidence_score": 0.95, "is_new_field_required": false, "inferred_data_type": "Text" }} ] }}
            """
            
            response = client.chat.completions.create(model="gpt-4o", messages=[{"role": "user", "content": prompt}], response_format={ "type": "json_object" })
            parsed_response = json.loads(response.choices[0].message.content)
            
            return {
                "message": "Successfully auto-mapped file headers to ISO Registry.", 
                "suggested_mappings": parsed_response.get("suggested_mappings", []), 
                "file_type": file_type,
                "headers": headers,
                "sample_row": sample_row
            }
        except Exception as e:
            raise ValueError(f"AI auto-mapping failed: {str(e)}")

    def extract_unstructured_payload(self, text_content: str, prompt_mappings: List[Dict[str, str]]) -> Dict[str, Any]:
        """
        Uses an LLM Agent to extract structured data from messy unstructured text (like PDFs) 
        based on a dynamic list of business prompts.
        """
        openai_api_key = os.getenv("OPENAI_API_KEY")
        if not openai_api_key:
            raise ValueError("OpenAI API key not configured. Cannot perform unstructured extraction.")

        try:
            client = OpenAI(api_key=openai_api_key)
            system_prompt = f"""
            You are an expert Data Extraction Agent. I am providing you with unstructured document text.
            Extract the data points according to the specific instructions provided.
            Return STRICTLY a JSON object where keys are the 'target_field' identifiers and values are the extracted strings/numbers.
            If a value cannot be found, return null for that key. Do not make up information.

            Extraction Instructions:
            {json.dumps(prompt_mappings, indent=2)}
            """
            
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"Document Text:\n{text_content[:100000]}"} # Safely bound text length
                ],
                response_format={ "type": "json_object" }
            )
            
            return json.loads(response.choices[0].message.content)
        except Exception as e:
            raise ValueError(f"AI unstructured extraction failed: {str(e)}")

    def run_scheduled_insights(self, db: Session) -> dict:
        """
        Finds and executes all scheduled insights that are due to run at the current time.
        This is the core logic for the scheduler job.
        """
        now = datetime.datetime.utcnow()
        insights_to_run = []

        # 1. Find all insights configured for a scheduled trigger
        scheduled_insights = db.query(models.InsightDefinition).filter(
            models.InsightDefinition.trigger_type == "SCHEDULED"
        ).all()

        # 2. Check which ones are due now
        for insight in scheduled_insights:
            cron_schedule = insight.trigger_config.get("cron")
            if cron_schedule and croniter.is_match(cron_schedule, now):
                insights_to_run.append(insight)
        
        if not insights_to_run:
            return {"executed_count": 0, "details": "No scheduled insights were due to run."}

        # 3. Execute the due insights via distributed Celery task dispatch
        from tasks import execute_insight_task
        for insight in insights_to_run:
            execute_insight_task.delay(insight.insight_id)

        return {"executed_count": len(insights_to_run), "details": "Dispatched to parallel Celery worker queue."}


class InsightsOrchestrator:
    """
    A dedicated, lightweight orchestrator for executing the analysis steps of an Insight.
    This is adapted from the main WorkflowExecutor to ensure consistent logic.
    """
    def __init__(self, db: Session):
        # Pre-load all necessary blueprints for execution, similar to WorkflowExecutor
        self.db = db
        cache = AssetCache(db)
        self.rule_sets_by_token_code = cache.rule_sets_by_token_code
        self.formulas_by_token_code = cache.formulas_by_token_code
        self.composite_formulas_by_token_code = cache.composite_formulas_by_token_code
        self.api_configs_by_id = cache.api_configs_by_id
        
        self.integration_dispatcher = IntegrationDispatcher(self.api_configs_by_id)

    def execute_steps(self, steps: List[Dict[str, Any]], context: Dict[str, Any]) -> (Dict[str, Any], List[str]):
        """
        Executes a list of orchestration steps against a given context.
        Returns the final context and a list of execution logs.
        """
        logs = []
        sorted_steps = sorted(steps, key=lambda s: s.get('sequence_number', 99))

        # Ensure api_responses dictionary exists in the context
        if 'api_responses' not in context:
            context['api_responses'] = {}

        for step in sorted_steps:
            should_invoke = True
            rule_token = step.get("invocation_rule_token")
            if rule_token:
                rule_set_model = self.rule_sets_by_token_code.get(rule_token)
                if rule_set_model:
                    bre = BusinessRuleEngine(rule_set_model.definition, None)
                    rule_passed, _, rule_logs = bre.execute(context)
                    logs.extend(rule_logs)
                    if not rule_passed:
                        should_invoke = False
                else:
                    should_invoke = False
                    logs.append(f"[WARN] Invocation rule '{rule_token}' not found. Skipping step.")

            if not should_invoke:
                continue

            try:
                step_type = step.get("step_type")
                target_token = step.get("target_token")

                if step_type == "BUSINESS_RULE":
                    rule_set_model = self.rule_sets_by_token_code.get(target_token)
                    if rule_set_model:
                        calc_engine = CalculationEngine(formula_library=self.formulas_by_token_code)
                        bre = BusinessRuleEngine(rule_set_model.definition, calc_engine)
                        _, context, rule_logs = bre.execute(context)
                        logs.extend(rule_logs)
                    else:
                        logs.append(f"[WARN] Business Rule Set '{target_token}' not found. Skipping.")

                elif step_type == "CALCULATION":
                    calc_engine = CalculationEngine(formula_library=self.formulas_by_token_code)
                    if target_token in self.formulas_by_token_code:
                        calc_result = calc_engine.execute_formula_by_token(target_token, context)
                        context = calc_result["final_context"]
                        logs.extend(calc_result["logs"])
                    elif target_token in self.composite_formulas_by_token_code:
                        composite = self.composite_formulas_by_token_code[target_token]
                        for comp_step in sorted(composite.steps, key=lambda s: s.sequence_number):
                            calc_result = calc_engine.execute_formula_by_token(comp_step.formula_token_code, context)
                            context = calc_result["final_context"]
                            logs.extend(calc_result["logs"])
                    else:
                        logs.append(f"[WARN] Calculation asset '{target_token}' not found. Skipping.")
                
                elif step_type == "API_CALL":
                    logs.append(f"Executing step: API_CALL '{target_token}'")
                    
                    success, api_logs, _ = self.integration_dispatcher.execute_api_call(
                        api_id=target_token,
                        context=context
                    )
                    logs.extend(api_logs)

                elif step_type == "EVENT_BROADCAST":
                    target_event_type = step.get("target_event_type")
                    if target_event_type:
                        logs.append(f"Broadcasting event: {target_event_type}")
                        trace_depth = context.get("__trace_depth__", 0) + 1
                        correlation_id = context.get("__correlation_id__", f"CORR-{uuid.uuid4().hex[:12]}")
                        asyncio.run(global_event_bus.broadcast(SystemEvent(
                            event_type=target_event_type,
                            source_context="InsightsOrchestrator",
                            payload=context,
                            trace_depth=trace_depth,
                            correlation_id=correlation_id
                        )))
            except Exception as e:
                logs.append(f"[ERROR] Insight analysis step failed: {str(e)}")
                # Continue to the next step

        return context, logs