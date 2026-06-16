import time
import csv
import io
import openpyxl
import datetime
import base64
import xml.etree.ElementTree as ET
from pypdf import PdfReader
from dbfread import DBF
import tempfile
import os

from database import RegionalSessionLocal
import models
from services.workflow_executor import WorkflowExecutor
from celery_app import celery_app

# This file represents tasks that would be executed by a distributed task queue worker (e.g., Celery).

@celery_app.task(name="process_file_task", bind=True, max_retries=3)
def process_file_task(self, job_id: str, mapper_id: str, workflow_id: str, file_contents_b64: str, filename: str, x_tenant_region: str = "DEFAULT", target_instance_id: str = None, document_type: str = None):
    """
    This function is executed by a dedicated background worker, completely isolated
    from the main web application process. It creates its own database session.
    """
    region_key = x_tenant_region.upper() if x_tenant_region else "DEFAULT"
    SessionClass = RegionalSessionLocal.get(region_key, RegionalSessionLocal["DEFAULT"])
    db = SessionClass()
    try:
        file_contents = base64.b64decode(file_contents_b64)

        job = db.query(models.IngestionJob).filter(models.IngestionJob.job_id == job_id).first()
        if not job or job.status != "PENDING":
            print(f"[WORKER_ERROR] Job {job_id} not found or not in PENDING state.")
            return

        job.status = "PROCESSING"
        job.processing_started_at = datetime.datetime.utcnow().isoformat()
        db.commit()

        mapper = db.query(models.PayloadMapperBlueprint).filter(models.PayloadMapperBlueprint.mapper_id == mapper_id).first()
        if not mapper:
            raise ValueError(f"Mapper with ID '{mapper_id}' not found.")

        prompt_mappings = [{"target_field": m.target_iso_field, "prompt": m.source_path} for m in mapper.mappings if m.reading_mode == "PROMPT"]
        records = []

        if prompt_mappings:
            # --- AGENTIC UNSTRUCTURED EXTRACTION FLOW (Universal) ---
            full_text = ""
            if filename.lower().endswith('.pdf'):
                reader = PdfReader(io.BytesIO(file_contents))
                full_text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
            elif filename.lower().endswith('.docx'):
                import docx
                doc = docx.Document(io.BytesIO(file_contents))
                full_text = "\n".join([para.text for para in doc.paragraphs])
            elif filename.lower().endswith(('.txt', '.csv', '.xml')):
                full_text = file_contents.decode('utf-8', errors='ignore')
            elif filename.lower().endswith(('.xlsx', '.xls')):
                workbook = openpyxl.load_workbook(io.BytesIO(file_contents), data_only=True)
                sheet = workbook.active
                for row in sheet.iter_rows(values_only=True):
                    full_text += ", ".join([str(c) for c in row if c is not None]) + "\n"
            else:
                # Universal Fallback
                full_text = file_contents.decode('utf-8', errors='ignore')
                
            from services.ai_services import AIService
            ai_service = AIService()
            extracted_data = ai_service.extract_unstructured_payload(full_text.strip(), prompt_mappings)
            records.append(extracted_data)

        else:
            # --- STANDARD STRUCTURED EXTRACTION FLOW ---
            if filename.lower().endswith('.csv'):
                decoded_file = file_contents.decode('utf-8')
                csv_reader = csv.DictReader(io.StringIO(decoded_file))
                records = [row for row in csv_reader]
            elif filename.lower().endswith(('.xlsx', '.xls')):
                workbook = openpyxl.load_workbook(io.BytesIO(file_contents))
                
                # 1. Extract CELL-based mappings (Global File Context / Summary Data)
                global_context = {}
                for mapping in [m for m in mapper.mappings if m.reading_mode == "CELL"]:
                    sheet = workbook[mapping.sheet_name] if mapping.sheet_name in workbook.sheetnames else workbook.active
                    val = sheet[mapping.cell_address].value if mapping.cell_address else None
                    global_context[mapping.source_path] = val
                    
                # 2. Extract COLUMN-based mappings (Transactional Rows)
                col_mappings = [m for m in mapper.mappings if m.reading_mode == "COLUMN"]
                sheet_name = col_mappings[0].sheet_name if col_mappings and col_mappings[0].sheet_name else workbook.active.title
                
                sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
                headers = [cell.value for cell in sheet[1]]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    record = dict(zip(headers, row))
                    record.update(global_context)
                    records.append(record)
                    
                # 3. Pre-Flight Control Totals Evaluation
                if mapper.file_control_totals:
                    for rule in mapper.file_control_totals:
                        sum_field = rule.get("sum_field")
                        target_cell_field = rule.get("target_cell_field")
                        calculated_sum = sum([float(r.get(sum_field, 0) or 0) for r in records])
                        expected_sum = float(global_context.get(target_cell_field, 0) or 0)
                        if abs(calculated_sum - expected_sum) > 0.01:
                            raise ValueError(f"Control Total Breach! Calculated sum of '{sum_field}' ({calculated_sum}) does not match expected total from '{target_cell_field}' ({expected_sum}).")
            elif filename.lower().endswith('.xml'):
                decoded_file = file_contents.decode('utf-8')
                root = ET.fromstring(decoded_file)
                for item in root:
                    record = {child.tag: child.text for child in item}
                    if record:
                        records.append(record)
            elif filename.lower().endswith('.pdf'):
                # Legacy fallback for PDFs without PROMPT mappings
                reader = PdfReader(io.BytesIO(file_contents))
                full_text = "\n".join([page.extract_text() for page in reader.pages if page.extract_text()])
                records.append({"raw_pdf_text": full_text.strip()})
            elif filename.lower().endswith('.dbf'):
                with tempfile.NamedTemporaryFile(delete=False, suffix='.dbf') as tmp:
                    tmp.write(file_contents)
                    tmp_path = tmp.name
                try:
                    for record in DBF(tmp_path, load=True):
                        records.append(dict(record))
                finally:
                    os.remove(tmp_path)
            else:
                raise ValueError(f"Unsupported file type for structured extraction: {filename}")

        job.total_records = len(records)
        db.commit()

        processed_count = 0

        if target_instance_id and document_type:
            # --- SCENARIO A: MULTI-STATEFUL CONVERGENCE ---
            # Target multiple paused workflow instances (e.g., a master rate sheet satisfying many checklists)
            instance_ids = [i.strip() for i in target_instance_id.split(',')]
            for inst_id in instance_ids:
                instance = db.query(models.WorkflowExecutionInstance).filter(models.WorkflowExecutionInstance.instance_id == inst_id).first()
                if instance:
                    transformed_records = [{m.target_iso_field: r.get(m.source_path, m.default_value) for m in mapper.mappings if m.target_iso_field} for r in records]
                    
                    merged_context = instance.current_context.copy()
                    merged_context[document_type] = transformed_records if len(transformed_records) > 1 else (transformed_records[0] if transformed_records else {})
                    
                    executor = WorkflowExecutor(db=db, workflow_id=instance.workflow_id)
                    executor.execute(initial_payload=merged_context, resume_from_node_id=instance.current_node_id, resume_trace=instance.execution_trace)
                    
                    instance.updated_at = datetime.datetime.utcnow().isoformat()
                    db.commit()
                    processed_count = max(processed_count, len(records))
        else:
            # --- SCENARIO B: MULTI-WORKFLOW FAN-OUT ---
            # Run multiple distinct workflow blueprints per row of the uploaded files
            target_workflows = [w.strip() for w in workflow_id.split(',')]
            for w_id in target_workflows:
                if w_id.lower() == 'resume': continue # Safety guard
                executor = WorkflowExecutor(db=db, workflow_id=w_id)
                for i, record in enumerate(records):
                    transformed_payload = {}
                    for mapping in mapper.mappings:
                        if record.get(mapping.source_path) is not None:
                            transformed_payload[mapping.target_iso_field] = record.get(mapping.source_path)
                    if transformed_payload:
                        executor.execute(initial_payload=transformed_payload)
                        processed_count += 1

        job.processed_records = processed_count
        job.status = "COMPLETED"
        job.completed_at = datetime.datetime.utcnow().isoformat()
        db.commit()

    except Exception as e:
        job.status = "FAILED"; job.error_message = str(e); job.completed_at = datetime.datetime.utcnow().isoformat(); db.commit()
    finally:
        db.close()

@celery_app.task(name="execute_insight_task", bind=True, max_retries=3)
def execute_insight_task(self, insight_id: str):
    """
    Isolated background worker process to execute an Insights Factory blueprint.
    Prevents heavy analyses from blocking the master scheduler thread.
    """
    from database import SessionLocal
    from services.ai_services import InsightsOrchestrator
    import models
    
    db = SessionLocal()
    try:
        insight = db.query(models.InsightDefinition).filter(models.InsightDefinition.insight_id == insight_id).first()
        if not insight: 
            return
            
        orchestrator = InsightsOrchestrator(db=db)
        orchestrator.execute_steps(insight.analysis_steps, {})
    except Exception as e:
        print(f"[WORKER_ERROR] Insight {insight_id} failed: {e}")
    finally:
        db.close()